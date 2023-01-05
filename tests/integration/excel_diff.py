import argparse
import os
import sys

from openpyxl import load_workbook

arg_parser = argparse.ArgumentParser()

arg_parser.add_argument("lhs_file", type=str, help="")

arg_parser.add_argument("rhs_file", type=str, help="")

args = arg_parser.parse_args()

if not os.path.exists(args.lhs_file):
    print(
        "error: path does not exist: {}".format(args.lhs_file), file=sys.stderr
    )
    exit(1)
if not os.path.exists(args.rhs_file):
    print(
        "error: path does not exist: {}".format(args.rhs_file), file=sys.stderr
    )
    exit(1)

lhs_wb = load_workbook(args.lhs_file)
rhs_wb = load_workbook(args.rhs_file)

lhs_sheet = lhs_wb.active
rhs_sheet = rhs_wb.active

if lhs_sheet.max_row != rhs_sheet.max_row:
    print("Excel files have different number of rows")
    exit(1)
if lhs_sheet.max_column != rhs_sheet.max_column:
    print("Excel files have different number of columns")
    exit(1)

errors = []
for row_num in range(1, lhs_sheet.max_row):
    for col_num in range(1, lhs_sheet.max_column):
        lhs_cell = lhs_sheet.cell(row=row_num, column=col_num)
        rhs_cell = rhs_sheet.cell(row=row_num, column=col_num)
        if lhs_cell.value != rhs_cell.value:
            errors.append(
                f"Cell {lhs_cell}: '{lhs_cell.value}' != '{rhs_cell.value}'"
            )

if errors:
    print("Excel Diff: files are not equal:")
    for error in errors:
        print(error)
    exit(1)
