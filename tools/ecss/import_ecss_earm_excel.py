import argparse
import os
import re
import shutil
from enum import Enum
from pathlib import Path
from typing import Dict, List

import openpyxl

from strictdoc import environment
from strictdoc.backend.sdoc.document_reference import DocumentReference
from strictdoc.backend.sdoc.models.document import SDocDocument
from strictdoc.backend.sdoc.models.document_config import DocumentConfig
from strictdoc.backend.sdoc.models.document_grammar import DocumentGrammar
from strictdoc.backend.sdoc.models.grammar_element import (
    GrammarElement,
    GrammarElementFieldString,
    GrammarElementFieldType,
)
from strictdoc.backend.sdoc.models.model import SDocDocumentIF
from strictdoc.backend.sdoc.models.node import SDocNode
from strictdoc.backend.sdoc.writer import SDWriter
from strictdoc.core.project_config import ProjectConfig


def strip_spaces_and_nbsp(input_string: str) -> str:
    """
    Strip leading and trailing space characters as well as &nbsp; symbols from a string.

    Some ECSS requirements have these unneeded characters. This method can be
    used to clean everything up.
    """
    assert isinstance(input_string, str), input_string
    return re.sub(
        r"(^[ \u00A0]+)|([ \u00A0]+$)", "", input_string, flags=re.MULTILINE
    )


class ECSS_EARM_Excel_Field(Enum):
    IE_PUID = "IE PUID"
    REQ_ID = "ECSS Req. Identifier"
    SOURCE_REFERENCE = "ECSS Source Reference"
    TYPE = "Type"
    STATUS = "ECSS Change Status"
    ORIGINAL_REQUIREMENT = "Original requirement"
    TEXT_OF_NOTE = "Text of Note of Original requirement"


ECSS_EARM_Excel_Field_Mapping: Dict[str, str] = {
    ECSS_EARM_Excel_Field.IE_PUID.value: "UID",
    ECSS_EARM_Excel_Field.REQ_ID.value: "ECSS_REQ_ID",
    ECSS_EARM_Excel_Field.SOURCE_REFERENCE.value: "ECSS_SOURCE",
    ECSS_EARM_Excel_Field.TYPE.value: "TYPE",
    ECSS_EARM_Excel_Field.STATUS.value: "STATUS",
    ECSS_EARM_Excel_Field.ORIGINAL_REQUIREMENT.value: "STATEMENT",
    ECSS_EARM_Excel_Field.TEXT_OF_NOTE.value: "COMMENT",
}


class ECSS_EARM_Excel_Importer:
    MAIN_SHEET = "DOORS 1.0 Current"

    @staticmethod
    def import_from_file(path_to_excel: str, path_to_output_dir: str):
        assert os.path.isfile(path_to_excel), path_to_excel

        workbook = openpyxl.load_workbook(path_to_excel, read_only=True)
        sheet = workbook[ECSS_EARM_Excel_Importer.MAIN_SHEET]
        if sheet is None:
            raise RuntimeError(
                "Couldn't open the main sheet of the workbook: "
                f"{ECSS_EARM_Excel_Importer.MAIN_SHEET}"
            )

        # The Excel file has the column names defined in the row #2.
        headers = [
            cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))
        ]

        for field_ in ECSS_EARM_Excel_Field_Mapping:
            assert field_ in headers, field_

        ecss_documents: Dict[str, SDocDocument] = {}

        for row in sheet.iter_rows(min_row=3):
            row_dict = {headers[i]: row[i].value for i in range(len(headers))}

            # Some ECSS requirements have no statements. Assuming that these
            # are deprecated or removed requirements. Skipping them entirely.
            requirement_statement = row_dict[
                ECSS_EARM_Excel_Field.ORIGINAL_REQUIREMENT.value
            ]
            if requirement_statement is None:
                continue

            ecss_document_id = row_dict[
                ECSS_EARM_Excel_Field.SOURCE_REFERENCE.value
            ]

            ecss_document: SDocDocument
            if ecss_document_id not in ecss_documents:
                ecss_document = SDocDocument(
                    mid=None,
                    title=ecss_document_id,
                    config=None,
                    view=None,
                    grammar=None,
                    section_contents=[],
                )
                ecss_document_config = DocumentConfig.default_config(
                    ecss_document
                )
                ecss_document_config.markup = "Text"

                ecss_document.config = ecss_document_config

                ecss_document.grammar = (
                    ECSS_EARM_Excel_Importer.create_ecss_grammar(ecss_document)
                )
                ecss_documents[ecss_document_id] = ecss_document
            else:
                ecss_document = ecss_documents[ecss_document_id]

            requirement_node = SDocNode(
                parent=ecss_document,
                node_type="REQUIREMENT",
                fields=[],
                relations=[],
                is_composite=False,
                section_contents=None,
            )
            requirement_node.ng_document_reference = DocumentReference()
            requirement_node.ng_document_reference.set_document(ecss_document)

            requirement_node.set_field_value(
                field_name=ECSS_EARM_Excel_Field_Mapping[
                    ECSS_EARM_Excel_Field.REQ_ID.value
                ],
                form_field_index=0,
                value=row_dict[ECSS_EARM_Excel_Field.REQ_ID.value],
            )
            requirement_node.set_field_value(
                field_name=ECSS_EARM_Excel_Field_Mapping[
                    ECSS_EARM_Excel_Field.IE_PUID.value
                ],
                form_field_index=0,
                value=row_dict[ECSS_EARM_Excel_Field.IE_PUID.value],
            )
            requirement_node.set_field_value(
                field_name=ECSS_EARM_Excel_Field_Mapping[
                    ECSS_EARM_Excel_Field.SOURCE_REFERENCE.value
                ],
                form_field_index=0,
                value=row_dict[ECSS_EARM_Excel_Field.SOURCE_REFERENCE.value],
            )
            requirement_node.set_field_value(
                field_name=ECSS_EARM_Excel_Field_Mapping[
                    ECSS_EARM_Excel_Field.TYPE.value
                ],
                form_field_index=0,
                value=row_dict[ECSS_EARM_Excel_Field.TYPE.value],
            )
            requirement_node.set_field_value(
                field_name=ECSS_EARM_Excel_Field_Mapping[
                    ECSS_EARM_Excel_Field.STATUS.value
                ],
                form_field_index=0,
                value=row_dict[ECSS_EARM_Excel_Field.STATUS.value],
            )
            requirement_node.set_field_value(
                field_name=ECSS_EARM_Excel_Field_Mapping[
                    ECSS_EARM_Excel_Field.ORIGINAL_REQUIREMENT.value
                ],
                form_field_index=0,
                value=strip_spaces_and_nbsp(
                    row_dict[ECSS_EARM_Excel_Field.ORIGINAL_REQUIREMENT.value]
                ),
            )

            if (
                requirement_note_ := row_dict[
                    ECSS_EARM_Excel_Field.TEXT_OF_NOTE.value
                ]
            ) is not None:
                requirement_node.set_field_value(
                    field_name=ECSS_EARM_Excel_Field_Mapping[
                        ECSS_EARM_Excel_Field.TEXT_OF_NOTE.value
                    ],
                    form_field_index=0,
                    value=strip_spaces_and_nbsp(requirement_note_),
                )

            ecss_document.section_contents.append(requirement_node)

        project_config = ProjectConfig.default_config(environment)

        ecss_output_dir = os.path.join(path_to_output_dir, "ecss")
        shutil.rmtree(ecss_output_dir, ignore_errors=True)
        Path(ecss_output_dir).mkdir(parents=True, exist_ok=True)

        for ecss_document_ in ecss_documents.values():
            sdoc_content = SDWriter(project_config).write(ecss_document_)
            document_file_name = f"{ecss_document_.reserved_title}.sdoc"
            document_path = os.path.join(ecss_output_dir, document_file_name)
            with open(document_path, "w") as output_file:
                output_file.write(sdoc_content)

    @staticmethod
    def create_ecss_grammar(document: SDocDocumentIF) -> DocumentGrammar:
        fields: List[GrammarElementFieldType] = []

        for (
            ecss_field_name_,
            sdoc_field_name_,
        ) in ECSS_EARM_Excel_Field_Mapping.items():
            fields.append(
                GrammarElementFieldString(
                    parent=None,
                    title=sdoc_field_name_,
                    human_title=ecss_field_name_,
                    required="False",
                )
            )

        requirement_element = GrammarElement(
            parent=None,
            tag="REQUIREMENT",
            property_is_composite="",
            property_prefix="",
            property_view_style="",
            fields=fields,
            relations=[],
        )

        elements: List[GrammarElement] = [requirement_element]
        grammar = DocumentGrammar(
            parent=document, elements=elements, import_from_file=None
        )
        grammar.is_default = False
        requirement_element.parent = grammar

        return grammar


def main():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "input_file", type=str, help="EARM_ECSS_export(DOORS-v1.0_Dec2024).xlsx"
    )
    parser.add_argument(
        "--output-dir", default="output/", type=str, help="TODO"
    )
    args = parser.parse_args()

    path_to_input_html = args.input_file
    if not os.path.isfile(path_to_input_html):
        print(f"not a file: {path_to_input_html}")  # noqa: T201
        exit(1)

    ECSS_EARM_Excel_Importer.import_from_file(
        path_to_input_html, args.output_dir
    )


if __name__ == "__main__":
    main()
