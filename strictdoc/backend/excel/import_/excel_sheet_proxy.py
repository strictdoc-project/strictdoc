"""
@relation(SDOC-SRS-152, scope=file)
"""

from enum import Enum
from typing import List

import openpyxl
import xlrd

from strictdoc.helpers.auto_described import auto_described
from strictdoc.helpers.cast import assert_cast


class ExcelLibType(Enum):
    XLRD = 1
    OPENPYXL = 2


@auto_described
class ExcelSheetProxy:
    """
    An adapter class to open xls(xlrd) or xlsx(openpyxl) with the same interface.
    """

    def __init__(self, file: str):
        self.file = file
        if file.endswith(".xlsx"):
            self.workbook = openpyxl.load_workbook(file)
            self.sheet = self.workbook.active
            if self.sheet is None:  # pragma: no cover
                raise RuntimeError(
                    "Couldn't open the first sheet of the workbook"
                )
            self.lib = ExcelLibType.OPENPYXL
        elif file.endswith(".xls"):
            self.workbook = xlrd.open_workbook(file, on_demand=True)
            self.sheet = self.workbook.sheet_by_index(0)
            self.lib = ExcelLibType.XLRD
        else:  # pragma: no cover
            raise AssertionError("Unsupported file format")

    @property
    def name(self) -> str:
        """
        Returns the name of the first sheet.
        """
        if self.lib == ExcelLibType.OPENPYXL:
            return str(self.sheet.title)
        elif self.lib == ExcelLibType.XLRD:
            return str(self.sheet.name)
        raise AssertionError

    @property
    def ncols(self) -> int:
        """
        The number of columns.
        """
        if self.lib == ExcelLibType.OPENPYXL:
            return assert_cast(self.sheet.max_column, int)
        elif self.lib == ExcelLibType.XLRD:
            return assert_cast(self.sheet.ncols, int)
        raise AssertionError

    @property
    def nrows(self) -> int:
        """
        The number of rows.
        """
        if self.lib == ExcelLibType.OPENPYXL:
            return assert_cast(self.sheet.max_row, int)
        elif self.lib == ExcelLibType.XLRD:
            return assert_cast(self.sheet.nrows, int)
        raise AssertionError

    def get_cell_value(self, row: int, col: int) -> str:
        """
        Returns the value at row/col as string.
        """
        if self.lib == ExcelLibType.OPENPYXL:
            return self.sheet.cell(row=row + 1, column=col + 1).value or ""
        elif self.lib == ExcelLibType.XLRD:
            return assert_cast(self.sheet.cell_value(row, col), str)
        raise AssertionError

    def row_values(self, row: int) -> List[str]:
        """
        Returns a full row as list.
        """
        if self.lib == ExcelLibType.OPENPYXL:
            return [(cell.value or "") for cell in self.sheet[row + 1]]
        elif self.lib == ExcelLibType.XLRD:
            return assert_cast(self.sheet.row_values(row), list)
        raise AssertionError
